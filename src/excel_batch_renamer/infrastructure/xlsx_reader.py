"""使用 openpyxl 读取固定表头、连续数据的任务表。"""

from pathlib import Path
from typing import Dict, List, Sequence

from openpyxl import load_workbook

from excel_batch_renamer.core.models import FolderTaskRow, ImageTaskRow


FOLDER_HEADERS = ("序号", "文件夹名称")
IMAGE_HEADERS = ("序号", "文件题名", "页次")
FOLDER_HEADER_ROW = 1
IMAGE_HEADER_ROW = 3


def list_worksheet_names(workbook_path: Path) -> List[str]:
    """读取工作簿的工作表名称，供图片任务 UI 选择。"""

    workbook = load_workbook(
        filename=str(workbook_path),
        read_only=True,
        data_only=True,
    )
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_folder_task(workbook_path: Path) -> List[FolderTaskRow]:
    """读取仅含一个工作表的文件夹任务表。"""

    workbook = load_workbook(
        filename=str(workbook_path),
        read_only=True,
        data_only=True,
    )
    try:
        if len(workbook.worksheets) != 1:
            raise ValueError("文件夹任务表必须且只能包含一个工作表")

        worksheet = workbook.worksheets[0]
        header_indexes = _read_header_indexes(
            worksheet,
            FOLDER_HEADERS,
            FOLDER_HEADER_ROW,
        )
        result = []
        for values in _iter_contiguous_rows(
            worksheet,
            header_indexes,
            FOLDER_HEADER_ROW + 1,
        ):
            result.append(
                FolderTaskRow(
                    sequence=_as_integer(values["序号"], "序号"),
                    folder_name=_as_optional_text(values["文件夹名称"]),
                )
            )
        return result
    finally:
        workbook.close()


def read_image_task(
    workbook_path: Path,
    worksheet_name: str,
) -> List[ImageTaskRow]:
    """读取图片任务表的指定工作表。"""

    workbook = load_workbook(
        filename=str(workbook_path),
        read_only=True,
        data_only=True,
    )
    try:
        if worksheet_name not in workbook.sheetnames:
            raise ValueError("工作表不存在：{}".format(worksheet_name))

        worksheet = workbook[worksheet_name]
        header_indexes = _read_header_indexes(
            worksheet,
            IMAGE_HEADERS,
            IMAGE_HEADER_ROW,
        )
        result = []
        for values in _iter_contiguous_rows(
            worksheet,
            header_indexes,
            IMAGE_HEADER_ROW + 1,
        ):
            result.append(
                ImageTaskRow(
                    file_title=_as_required_text(values["文件题名"], "文件题名"),
                    page_reference=_as_page_reference(values["页次"]),
                )
            )
        return result
    finally:
        workbook.close()


def _read_header_indexes(
    worksheet,
    required_headers: Sequence[str],
    header_row: int,
) -> Dict[str, int]:
    """读取指定行表头，忽略其中空白字符并返回必需列的零基索引。"""

    header_values = [
        _normalize_header(cell.value)
        for cell in next(
            worksheet.iter_rows(min_row=header_row, max_row=header_row)
        )
    ]
    row_label = {1: "第一行", 3: "第三行"}.get(
        header_row,
        "第{}行".format(header_row),
    )
    indexes = {}
    for header in required_headers:
        matches = [
            index for index, value in enumerate(header_values) if value == header
        ]
        if not matches:
            raise ValueError("{}缺少必需列：{}".format(row_label, header))
        if len(matches) > 1:
            raise ValueError("{}包含重复列：{}".format(row_label, header))
        indexes[header] = matches[0]
    return indexes


def _normalize_header(value) -> str:
    """删除表头中的全部 Unicode 空白字符，不修改数据单元格。"""

    if value is None:
        return ""
    return "".join(str(value).split())


def _iter_contiguous_rows(
    worksheet,
    header_indexes: Dict[str, int],
    first_data_row: int,
):
    """从指定行读取数据，在首个完全空白数据行结束。"""

    for row in worksheet.iter_rows(
        min_row=first_data_row,
        max_col=worksheet.max_column,
        values_only=True,
    ):
        if all(_is_blank(value) for value in row):
            break
        values = {
            header: row[index] for header, index in header_indexes.items()
        }
        yield values


def _is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _as_optional_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_required_text(value, field_name: str) -> str:
    text = _as_optional_text(value)
    if not text:
        raise ValueError("{}不能为空".format(field_name))
    return text


def _as_integer(value, field_name: str) -> int:
    if isinstance(value, bool):
        raise ValueError("{}必须是整数".format(field_name))
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)

    text = str(value).strip()
    if not text.isdigit():
        raise ValueError("{}必须是整数".format(field_name))
    return int(text)


def _as_page_reference(value) -> str:
    """保留范围写法，并把 Excel 数值页码规范成三位文本。"""

    if isinstance(value, bool):
        raise ValueError("页次格式无效")
    if isinstance(value, int):
        return "{:03d}".format(value)
    if isinstance(value, float) and value.is_integer():
        return "{:03d}".format(int(value))
    return _as_required_text(value, "页次")
