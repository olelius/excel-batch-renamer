import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from excel_batch_renamer.core.models import FolderTaskRow, ImageTaskRow
from excel_batch_renamer.infrastructure.xlsx_reader import (
    list_worksheet_names,
    read_folder_task,
    read_image_task,
)


class XlsxReaderTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.base_path = Path(self.temporary_directory.name)

    def tearDown(self):
        self.temporary_directory.cleanup()

    def test_read_folder_task_from_only_worksheet(self):
        path = self.base_path / "folders.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["序号", "文件夹名称"])
        worksheet.append([1, None])
        worksheet.append([2, "规划管理文件材料"])
        workbook.save(str(path))
        workbook.close()

        self.assertEqual(
            read_folder_task(path),
            [
                FolderTaskRow(sequence=1, folder_name=""),
                FolderTaskRow(sequence=2, folder_name="规划管理文件材料"),
            ],
        )

    def test_folder_task_rejects_multiple_worksheets(self):
        path = self.base_path / "multiple.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["序号", "文件夹名称"])
        workbook.create_sheet("第二页")
        workbook.save(str(path))
        workbook.close()

        with self.assertRaisesRegex(ValueError, "只能包含一个工作表"):
            read_folder_task(path)

    def test_folder_task_stops_at_first_completely_blank_row(self):
        path = self.base_path / "continuous.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["序号", "文件夹名称"])
        worksheet.append([1, "第一项"])
        worksheet.append([None, None])
        worksheet.append([2, "不会读取"])
        workbook.save(str(path))
        workbook.close()

        self.assertEqual(
            read_folder_task(path),
            [FolderTaskRow(sequence=1, folder_name="第一项")],
        )

    def test_read_selected_image_worksheet_and_ignore_other_columns(self):
        path = self.base_path / "images.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "1"
        worksheet.append(["备注", "序号", "页次", "文件题名"])
        worksheet.append(["忽略", 100, 1, "文件A"])
        worksheet.append(["忽略", 200, "005", "文件B"])
        worksheet.append(["忽略", 300, "010-017", "文件C"])
        worksheet.append([None, None, None, None])
        worksheet.append(["忽略", 400, "018-020", "不会读取"])
        workbook.create_sheet("2")
        workbook.save(str(path))
        workbook.close()

        self.assertEqual(list_worksheet_names(path), ["1", "2"])
        self.assertEqual(
            read_image_task(path, "1"),
            [
                ImageTaskRow(file_title="文件A", page_reference="001"),
                ImageTaskRow(file_title="文件B", page_reference="005"),
                ImageTaskRow(file_title="文件C", page_reference="010-017"),
            ],
        )

    def test_only_a_completely_blank_row_ends_reading(self):
        path = self.base_path / "not-completely-blank.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "1"
        worksheet.append(["序号", "文件题名", "页次", "备注"])
        worksheet.append([1, None, None, "该行并非完全空白"])
        workbook.save(str(path))
        workbook.close()

        with self.assertRaisesRegex(ValueError, "文件题名不能为空"):
            read_image_task(path, "1")

    def test_headers_must_be_in_first_row(self):
        path = self.base_path / "wrong-header.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["说明"])
        worksheet.append(["序号", "文件夹名称"])
        workbook.save(str(path))
        workbook.close()

        with self.assertRaisesRegex(ValueError, "第一行缺少必需列"):
            read_folder_task(path)

    def test_folder_headers_ignore_internal_unicode_whitespace(self):
        path = self.base_path / "folder-header-spaces.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append([" 序 号 ", "文件　夹\t名称"])
        worksheet.append([1, "规划 管理"])
        workbook.save(str(path))
        workbook.close()

        self.assertEqual(
            read_folder_task(path),
            [FolderTaskRow(sequence=1, folder_name="规划 管理")],
        )

    def test_image_headers_ignore_internal_unicode_whitespace(self):
        path = self.base_path / "image-header-spaces.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "1"
        worksheet.append(["序 号", "文 件　题 名", " 页\t次 "])
        worksheet.append([1, "文件 A", "001-001"])
        workbook.save(str(path))
        workbook.close()

        self.assertEqual(
            read_image_task(path, "1"),
            [ImageTaskRow(file_title="文件 A", page_reference="001-001")],
        )

    def test_normalized_duplicate_headers_are_rejected(self):
        path = self.base_path / "duplicate-normalized-header.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["序号", "序 号", "文件夹名称"])
        worksheet.append([1, 1, "名称"])
        workbook.save(str(path))
        workbook.close()

        with self.assertRaisesRegex(ValueError, "第一行包含重复列：序号"):
            read_folder_task(path)

    def test_image_task_requires_named_worksheet(self):
        path = self.base_path / "missing-sheet.xlsx"
        workbook = Workbook()
        workbook.active.title = "1"
        workbook.save(str(path))
        workbook.close()

        with self.assertRaisesRegex(ValueError, "工作表不存在"):
            read_image_task(path, "2")


if __name__ == "__main__":
    unittest.main()
